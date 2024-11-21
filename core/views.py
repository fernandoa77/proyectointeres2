from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from .forms import AmortizationForm, SinkingFundForm
import pandas as pd
from datetime import datetime, timedelta
from decimal import Decimal
import calendar
from django.contrib.auth.decorators import login_required
from .models import AmortizationTable, SinkingFund
from django.contrib.auth import login, logout
from django.contrib.auth.forms import UserCreationForm
from django.contrib import messages
import json
from datetime import date
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User

def generate_amortization_schedule(amount, interest_rate, term, payment_frequency, start_date):
    schedule = []
    rate_per_period = float(interest_rate) / (100 * payment_frequency)
    n_periods = term * payment_frequency
    
    # Cálculo del pago periódico
    payment = float(amount) * (rate_per_period * (1 + rate_per_period)**n_periods) / ((1 + rate_per_period)**n_periods - 1)
    
    remaining_balance = float(amount)
    current_date = start_date

    for period in range(1, int(n_periods) + 1):
        interest_payment = remaining_balance * rate_per_period
        principal_payment = payment - interest_payment
        remaining_balance -= principal_payment
        
        schedule.append({
            'month': period,
            'payment_date': current_date,
            'monthly_payment': payment,
            'interest': interest_payment,
            'principal': principal_payment,
            'remaining_balance': remaining_balance if remaining_balance > 0 else 0
        })
        
        # Actualizar la fecha según la frecuencia de pago
        if payment_frequency == 24:  # Quincenal
            current_date += timedelta(days=15)
        elif payment_frequency == 12:  # Mensual
            month_days = calendar.monthrange(current_date.year, current_date.month)[1]
            current_date += timedelta(days=month_days)
        elif payment_frequency == 2:  # Semestral
            for _ in range(6):
                month_days = calendar.monthrange(current_date.year, current_date.month)[1]
                current_date += timedelta(days=month_days)
        else:  # Anual
            for _ in range(12):
                month_days = calendar.monthrange(current_date.year, current_date.month)[1]
                current_date += timedelta(days=month_days)
    
    return schedule

def amortization_view(request):
    if request.method == 'POST':
        form = AmortizationForm(request.POST)
        if form.is_valid():
            # Obtener los datos del formulario
            amount = form.cleaned_data['amount']
            interest_rate = Decimal(str(form.cleaned_data['interest_rate']))
            term = form.cleaned_data['term']
            payment_frequency = int(form.cleaned_data['payment_frequency'])
            start_date = form.cleaned_data['start_date']

            # Generar el schedule
            schedule = generate_amortization_schedule(
                amount, 
                interest_rate, 
                term,
                payment_frequency,
                start_date
            )

            # Si se solicita la descarga
            if 'download' in request.POST:
                # Crear el DataFrame
                df = pd.DataFrame([{
                    'Período': entry['month'],
                    'Fecha de Pago': entry['payment_date'].strftime('%d/%m/%Y'),
                    'Pago': float(entry['monthly_payment']),
                    'Interés': float(entry['interest']),
                    'Principal': float(entry['principal']),
                    'Saldo Restante': float(entry['remaining_balance'])
                } for entry in schedule])

                # Crear el archivo Excel en memoria
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=tabla_amortizacion.xlsx'

                # Escribir el DataFrame al Excel
                with pd.ExcelWriter(response, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Amortización')
                    
                    # Ajustar anchos de columna
                    worksheet = writer.sheets['Amortización']
                    for idx, col in enumerate(df.columns):
                        max_length = max(
                            df[col].astype(str).apply(len).max(),
                            len(str(col))
                        )
                        worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_length + 4

                return response

            # Si no es descarga, continuar con el renderizado normal
            table_data_json = json.dumps({
                'amount': str(amount),
                'interest_rate': str(interest_rate),
                'term': term,
                'payment_frequency': payment_frequency,
                'start_date': start_date.strftime('%Y-%m-%d'),
                'schedule': [{
                    'month': entry['month'],
                    'payment_date': entry['payment_date'].strftime('%Y-%m-%d'),
                    'monthly_payment': float(entry['monthly_payment']),
                    'interest': float(entry['interest']),
                    'principal': float(entry['principal']),
                    'remaining_balance': float(entry['remaining_balance'])
                } for entry in schedule]
            })

            return render(request, 'amortization.html', {
                'form': form,
                'schedule': schedule,
                'table_data_json': table_data_json
            })

    else:
        form = AmortizationForm()

    return render(request, 'amortization.html', {'form': form})

def sinking_fund_view(request):
    if request.method == 'POST':
        form = SinkingFundForm(request.POST)
        if form.is_valid():
            # Obtener los datos del formulario
            target_amount = form.cleaned_data['target_amount']
            interest_rate = Decimal(str(form.cleaned_data['interest_rate']))
            term = form.cleaned_data['term']
            payment_frequency = int(form.cleaned_data['payment_frequency'])
            start_date = form.cleaned_data['start_date']

            # Generar el schedule
            schedule = generate_sinking_fund_schedule(
                target_amount, 
                interest_rate, 
                term,
                payment_frequency,
                start_date
            )

            # Si se solicita la descarga
            if 'download' in request.POST:
                # Crear el DataFrame
                df = pd.DataFrame([{
                    'Período': entry['period'],
                    'Fecha de Depósito': entry['payment_date'].strftime('%d/%m/%Y'),
                    'Depósito': float(entry['deposit']),
                    'Interés': float(entry['interest']),
                    'Saldo Acumulado': float(entry['balance'])
                } for entry in schedule])

                # Crear el archivo Excel en memoria
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=fondo_amortizacion.xlsx'

                # Escribir el DataFrame al Excel
                with pd.ExcelWriter(response, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Fondo de Amortización')
                    
                    # Ajustar anchos de columna
                    worksheet = writer.sheets['Fondo de Amortización']
                    for idx, col in enumerate(df.columns):
                        max_length = max(
                            df[col].astype(str).apply(len).max(),
                            len(str(col))
                        )
                        worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_length + 4

                return response

            # Si no es descarga, continuar con el renderizado normal
            fund_data = {
                'form_data': {
                    'target_amount': str(target_amount),
                    'interest_rate': str(interest_rate),
                    'term': term,
                    'payment_frequency': payment_frequency,
                    'start_date': start_date.strftime('%Y-%m-%d'),
                },
                'schedule': [{
                    'period': entry['period'],
                    'payment_date': entry['payment_date'].strftime('%Y-%m-%d'),
                    'deposit': float(entry['deposit']),
                    'interest': float(entry['interest']),
                    'balance': float(entry['balance'])
                } for entry in schedule]
            }
            fund_data_json = json.dumps(fund_data)

            return render(request, 'sinking_fund.html', {
                'form': form,
                'schedule': schedule,
                'fund_data_json': fund_data_json
            })

    else:
        form = SinkingFundForm()

    return render(request, 'sinking_fund.html', {'form': form})

def generate_sinking_fund_schedule(target_amount, interest_rate, term, payment_frequency, start_date):
    # Convertir todos los valores a Decimal
    target_amount = Decimal(str(target_amount))
    interest_rate = Decimal(str(interest_rate))
    term = Decimal(str(term))
    
    # Cálculo del depósito periódico
    rate_per_period = interest_rate / Decimal(str(payment_frequency)) / Decimal('100')
    n_periods = term * Decimal(str(payment_frequency))
    
    # Cálculo del denominador
    denominator = ((Decimal('1') + rate_per_period) ** n_periods - Decimal('1')) / rate_per_period
    deposit = target_amount / denominator

    schedule = []
    balance = Decimal('0')
    current_date = start_date

    for period in range(1, int(n_periods) + 1):
        interest = balance * rate_per_period
        balance += interest + deposit
        
        schedule.append({
            'period': period,
            'payment_date': current_date,
            'deposit': deposit,
            'interest': interest,
            'balance': balance,
        })
        
        # Incrementar la fecha según la frecuencia
        if payment_frequency == 24:  # Quincenal
            current_date += timedelta(days=15)
        elif payment_frequency == 12:  # Mensual
            month_days = calendar.monthrange(current_date.year, current_date.month)[1]
            current_date += timedelta(days=month_days)
        elif payment_frequency == 2:  # Semestral
            for _ in range(6):
                month_days = calendar.monthrange(current_date.year, current_date.month)[1]
                current_date += timedelta(days=month_days)
        else:  # Anual
            for _ in range(12):
                month_days = calendar.monthrange(current_date.year, current_date.month)[1]
                current_date += timedelta(days=month_days)

    return schedule

@login_required
def my_tables(request):
    tables = AmortizationTable.objects.filter(user=request.user)
    return render(request, 'my_tables.html', {'tables': tables})

@login_required
def my_funds(request):
    funds = SinkingFund.objects.filter(user=request.user)
    return render(request, 'my_funds.html', {'funds': funds})

@login_required
def save_amortization(request):
    if request.method == 'POST':
        table_data = request.POST.get('table_data')
        if table_data:
            AmortizationTable.objects.create(
                user=request.user,
                data=json.loads(table_data)
            )
            messages.success(request, 'Tabla guardada exitosamente.')
        return redirect('my_tables')
    return redirect('amortization')

@login_required
def save_fund(request):
    if request.method == 'POST':
        fund_data = request.POST.get('fund_data')
        if fund_data:
            SinkingFund.objects.create(
                user=request.user,
                data=json.loads(fund_data)
            )
            messages.success(request, 'Fondo guardado exitosamente.')
        return redirect('my_funds')
    return redirect('sinking_fund')

@login_required
def table_detail(request, id):
    table = get_object_or_404(AmortizationTable, id=id, user=request.user)
    return render(request, 'table_detail.html', {'table': table})

@login_required
def fund_detail(request, id):
    fund = get_object_or_404(SinkingFund, id=id, user=request.user)
    return render(request, 'fund_detail.html', {'fund': fund})

class CustomUserCreationForm(UserCreationForm):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Traducir los mensajes de error
        self.error_messages = {
            'password_mismatch': 'Las contraseñas no coinciden.',
            'password_too_short': 'La contraseña es demasiado corta.',
            'password_too_common': 'La contraseña es demasiado común.',
            'password_entirely_numeric': 'La contraseña no puede ser completamente numérica.',
        }
        # Traducir las etiquetas de los campos
        self.fields['username'].label = 'Nombre de usuario'
        self.fields['password1'].label = 'Contraseña'
        self.fields['password2'].label = 'Confirmar contraseña'
        # Traducir los mensajes de ayuda
        self.fields['username'].help_text = 'Requerido. 150 caracteres o menos. Letras, números y @/./+/-/_ solamente.'
        self.fields['password1'].help_text = 'Tu contraseña debe contener al menos 8 caracteres.'
        self.fields['password2'].help_text = 'Ingresa la misma contraseña que antes, para verificación.'

def signup(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, '¡Cuenta creada exitosamente! Bienvenido.')
            return redirect('home')
    else:
        form = CustomUserCreationForm()
    return render(request, 'registration/signup.html', {'form': form})

def home(request):
    """
    Vista para la página principal que muestra las opciones disponibles
    y un mensaje de bienvenida personalizado si el usuario está autenticado.
    """
    context = {
        'title': 'Bienvenido a AMORT',
        'description': 'Sistema de cálculo de tablas de amortización y fondos de amortización'
    }
    
    if request.user.is_authenticated:
        # Si el usuario está autenticado, agregar información personalizada
        tables_count = AmortizationTable.objects.filter(user=request.user).count()
        funds_count = SinkingFund.objects.filter(user=request.user).count()
        context.update({
            'tables_count': tables_count,
            'funds_count': funds_count
        })
    
    return render(request, 'home.html', context)

def logout_view(request):
    logout(request)
    messages.success(request, 'Has cerrado sesión exitosamente.')
    return redirect('home')

class DateEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.strftime('%Y-%m-%d')
        elif isinstance(obj, date):
            return obj.strftime('%Y-%m-%d')
        return super().default(obj)

def recursos_educativos(request):
    return render(request, 'recursos_educativos.html')

def acerca_de(request):
    return render(request, 'acerca_de.html')

def auto_login(request):
    # Si el usuario ya está autenticado, redirigir directamente a home
    if request.user.is_authenticated:
        return redirect('home')
    
    # Credenciales del usuario demo
    username = 'usuariodemo'
    password = 'demo142857'
    
    # Intentar obtener el usuario demo
    try:
        user = User.objects.get(username=username)
    except User.DoesNotExist:
        # Si no existe, crear el usuario demo
        user = User.objects.create_user(username=username, password=password)
    
    # Autenticar y hacer login
    user = authenticate(request, username=username, password=password)
    if user is not None:
        login(request, user)
        messages.success(request, 'Has iniciado sesión como usuario demo.')
    else:
        messages.error(request, 'Error al iniciar sesión como usuario demo.')
    
    return redirect('home')